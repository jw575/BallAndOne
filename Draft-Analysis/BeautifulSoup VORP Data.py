# Extract VORP, draft position, and salary data from basjetball-reference using Beautiful Soup
# Author: Jasper Wu
# Date:   06/09/2016

from bs4 import BeautifulSoup
from urllib.request import urlopen
from re import sub
from decimal import Decimal
import openpyxl
import re
import warnings
import pandas as pd

# Set ipython's max row display
pd.set_option('display.max_row', 1000)
# Set iPython's max column width to 50
pd.set_option('display.max_columns', 50)
warnings.filterwarnings("ignore")

# Gather data between these seasons
begDate = 2008
endDate = 2016

# Load data to this workbook. Change the location/name to whever you put the file
wbLocation = 'C:\\Users\Jasper\Documents\Basketball Blog Material\Posts\Jasper - Web Scraping\VORP Data.xlsx'
wb = openpyxl.load_workbook(wbLocation)

# We will dump our data into the sheet called 'VORP'
ws = wb.get_sheet_by_name('VORP')
last_row = ws.max_row

# Grab all advanced stats
def get_VORP(season, last_row):
    # Get advanced stats for all players for the given season and store them in Excel
    url = 'http://www.basketball-reference.com/leagues/NBA_' + str(season) + '_advanced.html'
    html = urlopen(url).read()
    soup = BeautifulSoup(html, 'lxml')
    rows = soup.find_all('tr', {'class': 'full_table'})
    tdcount = len(rows[1].findAll('td'))
    original_last = last_row

    # for every table row, get each TD
    for row in range(0, len(rows) - 1):
        last_row += 1
        columnCount = 3
        ws.cell(row=last_row, column=1).value = str(season-1) + "-" + str(season)[-2:]
        for td in range(1, tdcount):  # no need to start from 0 because 0 is Rk
            #Convert the data to numbers. If data can't be converted, pass it as a string
            try:
                ws.cell(row=last_row, column=columnCount).value = float(rows[row].findAll('td')[td].text)
            except ValueError:
                ws.cell(row=last_row,column=columnCount).value = rows[row].findAll('td')[td].text
            columnCount += 1

    # this part is for finding the player ID in a 'href'. Use regex to eliminate the noise.
    # From '/players/a/acyqu01.html', just return acyqu01
    p = re.compile('\/([^\/\?]+)\.')

    #Get the playerIDs from links
    for row in range(0, len(rows) - 1):
        original_last += 1
        columnCount = 2
        balls = rows[row].findAll('a')[0].get('href')
        ws.cell(row=original_last, column=columnCount).value = re.search(p, balls).group(1)
        # This is to calculate Wins Over Replacement Player (2.7 * VORP)
        columnCount = 31
        ws.cell(row=original_last, column=columnCount).value = float(ws.cell(row=original_last, column=columnCount-1).value) * 2.7

    wb.save(wbLocation)

# Returns dict of all row indexes where a unique player ID appears in the ws and seasons
# Keys: Season, rowIndex
def get_row_season(uniquePlayerID):
    row_season = dict.fromkeys(['Season','rowIndex'])
    seasons = []
    rows= []
    for row in range(2, ws.max_row+1):
        if ws.cell(row=row, column=2).value == uniquePlayerID:
            # Can't use "-" in seasons as dict key later on
            seasons.append(ws.cell(row=row, column=1).value.replace('-', '_'))
            rows.append(row)
    row_season['Season'] = seasons
    row_season['rowIndex'] = rows
    return row_season

# For players currently listed, get their draft/salary information and write it to the sheet.
def get_Draft_Salary(ws, uniquePlayerIDs):
    for player in uniquePlayerIDs:
        print(player)
        full_info = get_row_season(player)
        player_url = 'http://www.basketball-reference.com/players/' + str(player)[0] + '/' + str(player) + '.html'
        html = urlopen(player_url).read()
        soup = BeautifulSoup(html, 'lxml')
        salaryDict= {}

        #----------------------------- DRAFT Information ---------------------------------------
        #Older players in basketball-reference Draft data stored in person_image_offset. Newer players don't have
        # person_image_offset.
        if not soup.find_all('div', {'class': 'person_image_offset'}):
            draft = soup.find_all('div', {'class': 'margin_left_half'})[0]
        else:
            draft = soup.find_all('div', {'class': 'person_image_offset'})[0]

        if "Draft:" in draft.text:
            matchDraft = re.compile(r'[\n\r].*Draft:\s*([^\n\r]*)')
            draftInfo = re.search(matchDraft, draft.text).group(1)

            matchTeam = re.compile(r'([^,]+)')
            draftTeam = re.search(matchTeam, draftInfo).group(1)
            for row in full_info['rowIndex']:
                ws.cell(row=row, column=34).value = draftTeam

            matchPosition = re.compile(r'\((.*)\)')
            positionInfo = re.search(matchPosition, draftInfo).group(1)
            matchPosition2 = re.compile(r',\s(.*?)[a-z]')
            draftPosition = re.search(matchPosition2, positionInfo).group(1)
            for row in full_info['rowIndex']:
                ws.cell(row=row, column=33).value = int(draftPosition)

            matchYear = re.compile(r'\d{4}')
            draftYear = re.search(matchYear, draftInfo).group(0)
            for row in full_info['rowIndex']:
                ws.cell(row=row, column=32).value = int(draftYear)
        else:
            for row in full_info['rowIndex']:
                ws.cell(row=row, column=32).value = 'Undrafted'
                ws.cell(row=row, column=33).value = 'Undrafted'
                ws.cell(row=row, column=34).value = 'Undrafted'

        # ---------------------------- GET salary Information ---------------------------------------------
        # Get historical salaries first
        try:
            salariesPast = soup.find_all('div', {'id': 'div_salaries'})[0]
            for td in range(0, len(salariesPast.findAll('td')) - 4, 4):
                season = salariesPast.findAll('td')[td].text.replace('-', '_')
                money = Decimal(sub(r'[^\d.]', '', salariesPast.findAll('td')[td + 3].text))
                salaryDict[season] = money
        except:
            pass

        #APPEND current salary
        try:
            salariesNow = soup.find_all('div', {'id': 'all_contract'})[0]
            for entry in range(1, len(salariesNow.findAll('th'))):
                season = salariesNow.findAll('th')[entry].text.replace('-', '_')
                salary = Decimal(sub(r'[^\d.]', '', salariesNow.findAll('td')[entry].text))
                salaryDict[season] = salary
        except:
            pass

        counter = 0
        #Loop through all seasons the player played nd if match, get salary and write to Excel
        for playedSeason in full_info['Season']:
            for season, salary in salaryDict.items():
                if playedSeason == season:
                    ws.cell(row = full_info['rowIndex'][counter], column = 35).value = salary
                    counter += 1

    wb.save(wbLocation)

# Append data from time range to end of excel sheet
for year in range(begDate, endDate+1, 1):
    get_VORP(year, last_row)
    last_row = ws.max_row


# Get the unique playerIDs from the time range
VORP = pd.read_excel(wbLocation, 'VORP')
uniquePlayerIDs = VORP.PlayerID.unique()

# Append Draft data to all players
get_Draft_Salary(ws, uniquePlayerIDs)
