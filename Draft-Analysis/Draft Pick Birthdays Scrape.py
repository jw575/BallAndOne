'''
Pull the Position, Name, Position, Country and Birthday of every draft pick over a given time frame.
Write them to an excel.
Might be interesting also to calculate average age of drafted player when drafted.
Author: Jasper Wu
Date: 6/26/2016
'''

from bs4 import BeautifulSoup
from urllib.request import urlopen
import openpyxl
import re
import warnings
from datetime import datetime

warnings.filterwarnings("ignore")

# Use a 10 year time frame, with year = draft year
yearBeg = 1997
yearEnd = 2006

# Store our data in the 'Draft Data.xlsx' workbook
wbLocation = 'C:\\Users\Jasper\Documents\Basketball Blog Material\Posts\Jasper - Draft Age Effect\Draft Data.xlsx'
wb = openpyxl.load_workbook(wbLocation)

# We will write our data to the 'Draft Data Sheet'
ws = wb.get_sheet_by_name('Draft Data')

# Pull the draft data from Wikipedia
def get_draft_data(yearBeg, yearEnd, ws):
    # Pull the data for each Draft in our time range
    for year in range(yearBeg, yearEnd+1):
        print(year)
        url = 'https://en.wikipedia.org/wiki/' + str(year) + '_NBA_draft'
        html = urlopen(url).read()
        soup = BeautifulSoup(html, 'lxml')

        # From 2014 draft on, the class type of the Draft selections table changes to 'wikitable sortable
        # jquery-tablesorter'. Different for 2016 (uses plainrow headers)
        if year == 2004 or year == 2003:
            table = soup.find_all('table', {'class': 'wikitable sortable'})
        elif year < 2014:
            table = soup.find_all('table', {'class': 'wikitable sortable sortable'})
        elif year == 2016:
            table = soup.find_all('table', {'class': 'wikitable sortable plainrowheaders'})
        else:
            table = soup.find_all('table', {'class': 'wikitable sortable'})

        table = table[0]
        trCount = len(table.findAll('tr'))

        # Each row after after the first holds Draft information
        for row_ in range(1, trCount):
            last_row = ws.max_row+1
            section = table.findAll('tr')[row_]
            cells = section.findChildren('td')

            # Remember the playerLink for each player. We will need this to get birthdays
            playerLink = section.findAll('a')[0].get('href')
            playerName = section.findAll('a')[0].text

            # Store wiki link, year, player info
            ws.cell(row=last_row, column=1).value = playerLink
            ws.cell(row=last_row, column=2).value = year
            ws.cell(row=last_row, column=3).value = playerName

            # Begin storing all other information in column 4
            columnCount = 4

            # For every section, pull td text and ahref link. Don't pull player name since we already have it
            for cell_ in cells:
                # Convert the data to numbers. If data can't be converted, pass it as a string
                try:
                    ws.cell(row=last_row, column=columnCount).value = float(cell_.text.strip())
                except ValueError:
                    ws.cell(row=last_row, column=columnCount).value = cell_.text.strip()
                columnCount += 1

    wb.save(wbLocation)

# Get birthday data for each player
# Assuming player wikipedia link is in column 1, birthday in 11
def get_birthdays(ws):
    max_row = ws.max_row
    # For each player, get his birthday info from wikipedia link
    for eachRow_ in range(2, max_row+1):
        wikiLink = ws.cell(row=eachRow_, column=1).value
        wikiUrl = 'https://en.wikipedia.org' + str(wikiLink)
        html = urlopen(wikiUrl).read()
        soup = BeautifulSoup(html, 'lxml')
        try:
            table = soup.find_all('table', {'class': 'infobox vcard'})
            table = table[0]
            allRows = table.findAll('tr')

            # If in a league, tr 6 contains birthday information
            # If not in a league, 5th tr contains birthday information
            #print(wikiLink)
            if 'League' in allRows[3].text:
                birthdayBox = allRows[5]
            elif 'Personal' in allRows[1].text:
                birthdayBox = allRows[2]
            elif 'Personal' in allRows[0].text:
                birthdayBox = allRows[1]
            elif 'Personal' in allRows[2].text:
                birthdayBox = allRows[3]
            else:
                birthdayBox = allRows[4]
            birthdayCells = birthdayBox.findChildren('td')

            # Pull only the Birthday from this box
            # "\)([ ^\)]+)\(" Matches between ')' and '('
            matchRegex = re.compile(r'\)([^\)]+)\(')
            try:
                matched = re.search(matchRegex, birthdayCells[0].text)
                # Convert to date format
                birthday = matched.group(1).strip()
                try:
                    birthday = datetime.strptime(birthday, '%B %d, %Y').date()
                except ValueError:
                    birthday = datetime.strptime(birthday, '%d %B %Y').date()

                # Assign to birthday column for the given player's row
                ws.cell(row=eachRow_, column=11).value = birthday

            except:
                pass

        except:
            pass

    wb.save(wbLocation)

# Fill up Draft data
get_draft_data(yearBeg, yearEnd, ws)

# Get birthday information for all draft picks
get_birthdays(ws)
